# SK일렉링크_휴게소_충전기_현황.ver2.xlsx → sk_chargers.json (운영+설치예정) 생성. route/·docs/ 이중 저장
import sys, os, re, json
from collections import Counter
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)
XLSX = os.path.join(PARENT, "SK일렉링크_휴게소_충전기_현황.ver3_주말혼잡도.xlsx")
SHEET = "SK일렉링크 휴게소 충전기"
OUTS = [os.path.join(HERE, "sk_chargers.json"), os.path.join(PARENT, "sk_chargers.json")]

# 설치예정(좌표 없는) 휴게소는 주소를 지오코딩 — build_geocode.geocode 재사용(.env의 KAKAO_REST_KEY)
sys.path.insert(0, PARENT)
import build_geocode as geo

# (kW, 엑셀 열 인덱스, 급속/완속)
KW_COLS = [(350, 9, "급속"), (200, 10, "급속"), (100, 11, "급속"), (50, 12, "급속"), (7, 13, "완속")]
# ver3: 등급·경쟁강도 삭제 → 응모권=20, 혼잡도(주말)=21~25, 충전단가=31
C = dict(no=0, name=1, region=2, addr=3, lat=4, lng=5, total=6, conn=14, maker=15,
         year=16, lucky=17, status=18, menu=19, ticket=20, price=31)
CONG_SLOTS = [("00-08", 21), ("08-12", 22), ("12-16", 23), ("16-20", 24), ("20-24", 25)]


def clean(s):
    return s.encode("utf-8", "ignore").decode("utf-8") if isinstance(s, str) else s


def splist(v):  # "A, B" → ["A","B"] (커넥터/제조사/설치년도)
    return sorted({p.strip() for p in re.split(r"\s*,\s*", str(v)) if p.strip()}) if v else []


def parse_lucky(v):  # "4/4" → 4 (적용 대수)
    m = re.match(r"\s*(\d+)", str(v or ""))
    return int(m.group(1)) if m else 0


def parse_status(v, total):  # "정상사용 4" → {"정상사용":4}
    s = str(v or "").strip()
    if not s:
        return {}
    m = re.match(r"(.+?)\s*(\d+)\s*$", s)
    return {m.group(1).strip(): int(m.group(2))} if m else {s: total}


def parse_price(v):  # 347.2 / 391 / "295원" → 숫자(정수면 int)
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        p = float(v)
    else:
        m = re.search(r"[\d.]+", str(v))
        if not m:
            return None
        p = float(m.group())
    return int(p) if p == int(p) else p


wb = load_workbook(XLSX, data_only=True)
ws = wb[SHEET]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[C["name"]]]

stations = []
planned_ok = 0
planned_fail = []
kw_mismatch = []
for r in rows:
    name = clean(r[C["name"]]).strip()
    region = (clean(r[C["region"]]) or "").strip()
    addr = (clean(r[C["addr"]]) or "").strip()
    ticket = int(r[C["ticket"]]) if isinstance(r[C["ticket"]], (int, float)) else None
    lat, lng = r[C["lat"]], r[C["lng"]]
    has_coord = isinstance(lat, (int, float)) and isinstance(lng, (int, float))

    if has_coord:  # 운영 충전소
        by = []
        for kw, col, sp in KW_COLS:
            v = r[col]
            if isinstance(v, (int, float)) and v > 0:
                by.append({"kw": kw, "speed": sp, "count": int(v)})
        total = int(r[C["total"]]) if isinstance(r[C["total"]], (int, float)) else sum(b["count"] for b in by)
        if by and sum(b["count"] for b in by) != total:
            kw_mismatch.append((name, total, sum(b["count"] for b in by)))
        stations.append({
            "name": name, "lat": round(float(lat), 7), "lng": round(float(lng), 7),
            "address": addr, "region": region, "total": total,
            "fast": sum(b["count"] for b in by if b["speed"] == "급속"),
            "slow": sum(b["count"] for b in by if b["speed"] == "완속"),
            "byKw": by, "makers": splist(r[C["maker"]]),
            "years": splist(r[C["year"]]), "luckypass": parse_lucky(r[C["lucky"]]),
            "status": parse_status(r[C["status"]], total),
            "tickets": ticket, "planned": False,
            "price": parse_price(r[C["price"]]),
            "congestion": [{"t": lab, "lv": r[col]} for lab, col in CONG_SLOTS if r[col] not in (None, "")],
        })
    else:  # 설치예정 — 주소 지오코딩
        res = geo.geocode(name, addr)
        if not res:
            planned_fail.append(name)
            continue
        glat, glng, method, _ = res
        planned_ok += 1
        stations.append({  # 설치예정: 응모권·혼잡도·단가 없음
            "name": name, "lat": round(float(glat), 7), "lng": round(float(glng), 7),
            "address": addr, "region": region, "planned": True,
            "approx": method.startswith("approx:"),
        })

stations.sort(key=lambda s: s["name"])
data = json.dumps(stations, ensure_ascii=False, indent=1)
for out in OUTS:
    open(out, "w", encoding="utf-8").write(data)

# ---- 리포트 ----
op = [s for s in stations if not s["planned"]]
pl = [s for s in stations if s["planned"]]
print(f"[생성] {len(stations)}곳 → route/, docs/  (운영 {len(op)} · 설치예정 {len(pl)})")
print(f"운영 충전기 합계: {sum(s['total'] for s in op)}대")
print(f"설치예정 지오코딩: 성공 {planned_ok} · 실패 {len(planned_fail)}")
if planned_fail:
    print("  [지오코딩 실패]", planned_fail)
approx = [s["name"] for s in pl if s.get("approx")]
if approx:
    print(f"  근사좌표(읍·면 단위) {len(approx)}곳:", approx)
print("응모권 분포(운영):", dict(Counter(s["tickets"] for s in op)))
if kw_mismatch:
    print("총충전기 ≠ kW합 불일치:", kw_mismatch)
